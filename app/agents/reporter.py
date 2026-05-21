"""
Report Generator
Produces a styled Excel control sheet from a BatchResult.
Green/Amber/Red cell fills per field, issues log, works reconciliation.
"""
import io
from datetime import datetime
from app.schemas.models import BatchResult, RAGStatus

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# Colour palette
COLOURS = {
    RAGStatus.GREEN:   {"fill": "1A4731", "font": "22C55E"},
    RAGStatus.AMBER:   {"fill": "451A03", "font": "F59E0B"},
    RAGStatus.RED:     {"fill": "450A0A", "font": "EF4444"},
    RAGStatus.MISSING: {"fill": "1C2320", "font": "4A5C54"},
    "header":          {"fill": "0C0F0E", "font": "22C55E"},
    "subheader":       {"fill": "141918", "font": "6B7C74"},
    "row_odd":         {"fill": "141918", "font": "E5EDE8"},
    "row_even":        {"fill": "1C2320", "font": "E5EDE8"},
    "title":           {"fill": "0C0F0E", "font": "E5EDE8"},
}


def hex_fill(hex_str: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_str)


def hex_font(hex_str: str, bold=False, size=10) -> Font:
    return Font(color=hex_str, bold=bold, size=size, name="Courier New")


def thin_border() -> Border:
    s = Side(style="thin", color="2A3330")
    return Border(left=s, right=s, top=s, bottom=s)


def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def generate_excel(result: BatchResult) -> bytes:
    """Generate styled Excel control sheet from BatchResult."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl not installed")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Control Sheet ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Control Sheet"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    # Row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 16

    # Column widths
    col_widths = {
        "A": 28,  # Field
        "B": 8,   # Priority
        "C": 30,  # Submission
        "D": 28,  # Valuation
        "E": 28,  # Survey
        "F": 28,  # Questionnaire
        "G": 20,  # Works
        "H": 10,  # RAG
        "I": 35,  # Note
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Title block
    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = f"IRISH HOMES — MTR BATCH REVIEW CONTROL SHEET"
    title.fill = hex_fill("0C0F0E")
    title.font = Font(color="22C55E", bold=True, size=13, name="Courier New")
    title.alignment = center()

    ws.merge_cells("A2:I2")
    sub = ws["A2"]
    sub.value = f"Property: {result.address}  |  Ref: {result.property_ref}  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    sub.fill = hex_fill("141918")
    sub.font = Font(color="6B7C74", size=9, name="Courier New")
    sub.alignment = center()

    # Status summary row
    ws.merge_cells("A3:C3")
    status_cell = ws["A3"]
    status_colours = {
        RAGStatus.RED: ("450A0A", "EF4444"),
        RAGStatus.AMBER: ("451A03", "F59E0B"),
        RAGStatus.GREEN: ("1A4731", "22C55E"),
    }
    sc = status_colours[result.overall_status]
    status_cell.value = f"OVERALL: {result.overall_status.value} — {result.red_count} RED  {result.amber_count} AMBER  {result.green_count} GREEN"
    status_cell.fill = hex_fill(sc[0])
    status_cell.font = Font(color=sc[1], bold=True, size=10, name="Courier New")
    status_cell.alignment = center()

    # Doc presence row
    doc_cells = {"D3": "submission", "E3": "valuation", "F3": "survey", "G3": "questionnaire", "H3": "works"}
    for cell_addr, doc_type in doc_cells.items():
        cell = ws[cell_addr]
        present = result.doc_summary.get(doc_type, False)
        cell.value = f"{'✓' if present else '✗'} {doc_type.upper()}"
        cell.fill = hex_fill("1A4731" if present else "450A0A")
        cell.font = Font(color="22C55E" if present else "EF4444", size=8, name="Courier New")
        cell.alignment = center()

    # Merge I3
    ws["I3"].value = f"Docs: {sum(result.doc_summary.values())}/5"
    ws["I3"].fill = hex_fill("141918")
    ws["I3"].font = Font(color="6B7C74", size=9, name="Courier New")
    ws["I3"].alignment = center()

    # Column headers
    headers = ["FIELD", "PRI", "SUBMISSION", "VALUATION", "SURVEY", "QUESTIONNAIRE", "WORKS", "RAG", "NOTE"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = hex_fill("1C2320")
        cell.font = Font(color="4A5C54", bold=True, size=8, name="Courier New")
        cell.alignment = center()
        cell.border = thin_border()

    # Data rows
    for row_idx, field in enumerate(result.fields, 5):
        ws.row_dimensions[row_idx].height = 28
        bg = "141918" if row_idx % 2 == 0 else "1C2320"
        rag = field.status

        row_data = [
            field.field,
            field.priority[:4],  # CRIT / HIGH / MED
            field.submission or "—",
            field.valuation or "—",
            field.survey or "—",
            field.questionnaire or "—",
            field.works or "—",
            rag.value,
            field.note or "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value)[:200] if value else "")
            cell.border = thin_border()

            # RAG cell gets colour fill
            if col_idx == 8:
                c = COLOURS[rag]
                cell.fill = hex_fill(c["fill"])
                cell.font = Font(color=c["font"], bold=True, size=9, name="Courier New")
                cell.alignment = center()
            elif col_idx in (3, 4, 5, 6, 7):
                # Value cells — light colour from RAG
                cell.fill = hex_fill(bg)
                cell.font = Font(color=COLOURS[rag]["font"], size=9, name="Courier New")
                cell.alignment = left()
            elif col_idx == 9:
                # Note cell
                cell.fill = hex_fill(bg)
                cell.font = Font(color="F59E0B" if rag in (RAGStatus.RED, RAGStatus.AMBER) else "6B7C74",
                                size=8, name="Courier New", italic=True)
                cell.alignment = left()
            else:
                cell.fill = hex_fill(bg)
                cell.font = Font(color="E5EDE8" if col_idx == 1 else "6B7C74",
                                size=9, name="Courier New",
                                bold=(col_idx == 1))
                cell.alignment = left()

    # ── Sheet 2: Issues Log ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Issues Log")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 60
    ws2.column_dimensions["D"].width = 50

    ws2.merge_cells("A1:D1")
    ws2["A1"].value = f"ISSUES LOG — {result.address}"
    ws2["A1"].fill = hex_fill("0C0F0E")
    ws2["A1"].font = Font(color="22C55E", bold=True, size=12, name="Courier New")
    ws2["A1"].alignment = center()
    ws2.row_dimensions[1].height = 28

    for col, hdr in zip(["A", "B", "C", "D"], ["SEV", "TITLE", "DESCRIPTION", "SOURCE"]):
        cell = ws2[f"{col}2"]
        cell.value = hdr
        cell.fill = hex_fill("1C2320")
        cell.font = Font(color="4A5C54", bold=True, size=8, name="Courier New")
        cell.alignment = center()
        cell.border = thin_border()

    for i, issue in enumerate(result.issues, 3):
        ws2.row_dimensions[i].height = 40
        bg = "141918" if i % 2 == 0 else "1C2320"
        c = COLOURS[issue.severity]

        ws2.cell(i, 1, issue.severity.value).fill = hex_fill(c["fill"])
        ws2.cell(i, 1).font = Font(color=c["font"], bold=True, size=9, name="Courier New")
        ws2.cell(i, 1).alignment = center()
        ws2.cell(i, 1).border = thin_border()

        for col, val in zip([2, 3, 4], [issue.title, issue.description, issue.source]):
            cell = ws2.cell(i, col, str(val)[:500])
            cell.fill = hex_fill(bg)
            cell.font = Font(color="E5EDE8" if col == 2 else "A0B0A8", size=9, name="Courier New")
            cell.alignment = left()
            cell.border = thin_border()

    # ── Sheet 3: Works Reconciliation ─────────────────────────────────────────
    ws3 = wb.create_sheet("Works Reconciliation")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 80
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 10

    ws3.merge_cells("A1:E1")
    ws3["A1"].value = f"WORKS RECONCILIATION — {result.address}"
    ws3["A1"].fill = hex_fill("0C0F0E")
    ws3["A1"].font = Font(color="22C55E", bold=True, size=12, name="Courier New")
    ws3["A1"].alignment = center()

    for col, hdr in zip(["A", "B", "C", "D", "E"], ["#", "WORK DESCRIPTION", "IN SURVEY", "IN WORKS LIST", "STATUS"]):
        cell = ws3[f"{col}2"]
        cell.value = hdr
        cell.fill = hex_fill("1C2320")
        cell.font = Font(color="4A5C54", bold=True, size=8, name="Courier New")
        cell.alignment = center()

    for i, item in enumerate(result.works_reconciliation, 3):
        ws3.row_dimensions[i].height = 28
        bg = "141918" if i % 2 == 0 else "1C2320"
        c = COLOURS[item.status]

        ws3.cell(i, 1, item.number).fill = hex_fill(bg)
        ws3.cell(i, 1).font = Font(color="6B7C74", size=9, name="Courier New")
        ws3.cell(i, 1).alignment = center()

        ws3.cell(i, 2, item.description).fill = hex_fill(bg)
        ws3.cell(i, 2).font = Font(color="E5EDE8", size=9, name="Courier New")
        ws3.cell(i, 2).alignment = left()

        ws3.cell(i, 3, "✓" if item.in_survey else "✗").fill = hex_fill(bg)
        ws3.cell(i, 3).font = Font(color="22C55E" if item.in_survey else "EF4444", size=11, name="Courier New")
        ws3.cell(i, 3).alignment = center()

        ws3.cell(i, 4, "✓" if item.in_works else "✗").fill = hex_fill(bg)
        ws3.cell(i, 4).font = Font(color="22C55E" if item.in_works else "EF4444", size=11, name="Courier New")
        ws3.cell(i, 4).alignment = center()

        ws3.cell(i, 5, item.status.value).fill = hex_fill(c["fill"])
        ws3.cell(i, 5).font = Font(color=c["font"], bold=True, size=9, name="Courier New")
        ws3.cell(i, 5).alignment = center()

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
