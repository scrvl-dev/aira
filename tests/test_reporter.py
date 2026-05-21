import io

import openpyxl

from app.agents.reconciler import reconcile
from app.agents.reporter import generate_excel


def test_generate_excel_produces_valid_workbook(clean_models):
    result = reconcile(clean_models, "BATCH-XLSX")
    data = generate_excel(result)
    assert isinstance(data, bytes)
    assert len(data) > 0

    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Control Sheet", "Issues Log", "Works Reconciliation"]


def test_control_sheet_lists_all_fields(clean_models):
    result = reconcile(clean_models, "BATCH-XLSX")
    wb = openpyxl.load_workbook(io.BytesIO(generate_excel(result)))
    ws = wb["Control Sheet"]
    # Header row is row 4, data starts row 5
    field_cells = [ws.cell(row=r, column=1).value for r in range(5, 5 + len(result.fields))]
    assert "Open Market Value" in field_cells
    assert "Eircode" in field_cells
